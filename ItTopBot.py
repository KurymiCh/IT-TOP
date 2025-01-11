import openpyxl
import telebot
from openpyxl import load_workbook
import xlrd

TOKEN = '7860264700:AAE44AGv-IzP3m4v2C5O2kxQtfh6wKaFyZU'
ItBot = telebot.TeleBot(TOKEN)

@ItBot.message_handler(commands=['start', 'go'])
def start(message):
    ItBot.send_message(message.chat.id, "Айти бот готов к работе! "
                                        "Отправьте мне Excel файл с отчетом по домашним заданиям.")


@ItBot.message_handler(content_types=['document'])
def handle_document(message):
    document = message.document

    file_name = document.file_name

    if not file_name.endswith('.xlsx'):
        ItBot.send_message(message.chat.id, "Ошибка: файл должен иметь расширение .xlsx!")
        return

    try:
        file_info = ItBot.get_file(document.file_id)
        downloaded_file = ItBot.download_file(file_info.file_path)
        with open(file_name, 'wb') as new_file:
            new_file.write(downloaded_file)
        ItBot.send_message(message.chat.id, "Файл успешно загружен!")

    except Exception as e:
        ItBot.send_message(message.chat.id, f"Произошла ошибка при загрузке файла: {str(e)}")


@ItBot.message_handler(commands=['procentage_home_work'])
def handle_text(message):

    file_name = 'Отчет по домашним заданиям.xlsx'
    workbook = load_workbook(file_name)
    sheet = workbook.active

    @ItBot.message_handler()
    def calc(message):
        total_homework = 0
        verified_homework = 0
        fio_to_search = message.text

        for row in sheet.iter_rows(min_row=2, values_only=True):
            current_fio = row[1]
            if current_fio == fio_to_search:
                total_homework = row[4]
                verified_homework = row[5]
                break

        if total_homework > 0:
            procentage = (verified_homework / total_homework) * 100
            ItBot.send_message(message.chat.id, f'{int(procentage)}% проверенных дз у преподавателя {current_fio}')
        else:
            ItBot.send_message(message.chat.id, f'ФИО преподавателя не найдено или нет полученных домашних заданий')


@ItBot.message_handler(commands=['att'])
def attendence(message):
    file_name = 'Отчет по посещаемости студентов.xlsx'
    workbook = load_workbook(file_name)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=3, values_only=True):
            fio = row[0]
            if row[1] is not None:
                attendance = int(row[4].replace('%', '').strip())
            else:
                continue
            if attendance < 65:
                ItBot.send_message(message.chat.id, f"{fio}, Посещаемость: {attendance}%")



ItBot.polling(none_stop=True)